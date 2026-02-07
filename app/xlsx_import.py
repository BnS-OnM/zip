"""
Module for importing data from Excel files to Odoo.
Handles Product (product.template) and Sale Order (sale.order) Excel files.
"""
import openpyxl
import pandas as pd
import re
from io import BytesIO
from typing import List, Dict, Any, Optional
import logging

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

# Constants
DEFAULT_VAT_PERCENT = 21  # Default VAT rate for Belgium

# Module-level product catalog storage
_PRODUCT_CATALOG: List[Dict[str, Any]] = []


def parse_product_xlsx(xlsx_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Parse Product (product.template).xlsx file to extract product data.
    
    Expected columns:
    - default_code (Internal Reference / ArtikelNr)
    - name (Product Name)
    - list_price (Sales Price)
    - standard_price (Cost)
    - type (Product Type)
    - categ_id (Product Category)
    
    Args:
        xlsx_bytes: Bytes content of the Excel file
    
    Returns:
        List of product dictionaries with keys matching Odoo product.template fields
    """
    products = []
    
    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        logger.info(f"Product Excel headers: {headers}")
        
        # Map common Excel column names to Odoo field names
        field_mapping = {
            'artikelnr': 'default_code',
            'artikel': 'default_code',
            'default_code': 'default_code',
            'internal reference': 'default_code',
            'naam': 'name',
            'name': 'name',
            'product name': 'name',
            'verkoopprijs': 'list_price',
            'sales price': 'list_price',
            'list_price': 'list_price',
            'kostprijs': 'standard_price',
            'cost': 'standard_price',
            'standard_price': 'standard_price',
            'type': 'type',
            'product type': 'type',
        }
        
        # Create column index mapping
        column_mapping = {}
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if header_lower in field_mapping:
                    column_mapping[field_mapping[header_lower]] = idx
        
        logger.info(f"Column mapping: {column_mapping}")
        
        # Parse data rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            product = {}
            
            # Extract mapped fields
            for field_name, col_idx in column_mapping.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        product[field_name] = value
            
            # Skip if no default_code (required field)
            if 'default_code' not in product or not product['default_code']:
                logger.warning(f"Skipping row {row_idx}: No default_code found")
                continue
            
            # Ensure name is set (use default_code if not available)
            if 'name' not in product or not product['name']:
                product['name'] = product['default_code']
            
            products.append(product)
        
        logger.info(f"Parsed {len(products)} products from Excel file")
        
    except Exception as e:
        logger.error(f"Error parsing product Excel file: {str(e)}")
        raise
    
    return products


def parse_sale_order_xlsx(xlsx_bytes: bytes) -> Dict[str, Any]:
    """
    Parse Verkooporder (sale.order).xlsx file to extract sale order data.
    
    Expected structure:
    - Order header information (partner, date, etc.)
    - Order lines with product codes, descriptions, quantities, prices
    
    Args:
        xlsx_bytes: Bytes content of the Excel file
    
    Returns:
        Dictionary with order header data and list of order lines
    """
    order_data = {
        'lines': []
    }
    
    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        logger.info(f"Sale Order Excel headers: {headers}")
        
        # Map common Excel column names to order line fields
        field_mapping = {
            'artikelnr': 'product_code',
            'artikel': 'product_code',
            'default_code': 'product_code',
            'product code': 'product_code',
            'omschrijving': 'description',
            'description': 'description',
            'name': 'description',
            'hoeveelheid': 'quantity',
            'quantity': 'quantity',
            'qty': 'quantity',
            'eenheidsprijs': 'unit_price',
            'unit price': 'unit_price',
            'price unit': 'unit_price',
            'price_unit': 'unit_price',
            'btw': 'tax_percent',
            'tax': 'tax_percent',
            'vat': 'tax_percent',
        }
        
        # Create column index mapping
        column_mapping = {}
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if header_lower in field_mapping:
                    column_mapping[field_mapping[header_lower]] = idx
        
        logger.info(f"Column mapping: {column_mapping}")
        
        # Parse data rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            line = {}
            
            # Extract mapped fields
            for field_name, col_idx in column_mapping.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        line[field_name] = value
            
            # Skip if no description and no product code
            if 'description' not in line and 'product_code' not in line:
                logger.warning(f"Skipping row {row_idx}: No description or product code")
                continue
            
            # Ensure description is set
            if 'description' not in line or not line['description']:
                line['description'] = line.get('product_code', 'Product')
            
            # Convert types and set defaults
            try:
                line['quantity'] = int(line['quantity']) if line.get('quantity') else 1
                line['unit_price'] = float(line['unit_price']) if line.get('unit_price') else 0.0
                line['tax_percent'] = int(line['tax_percent']) if line.get('tax_percent') else DEFAULT_VAT_PERCENT
            except (ValueError, TypeError) as e:
                logger.warning(f"Error converting values in row {row_idx}: {e}")
                continue
            
            order_data['lines'].append(line)
        
        logger.info(f"Parsed {len(order_data['lines'])} order lines from Excel file")
        
    except Exception as e:
        logger.error(f"Error parsing sale order Excel file: {str(e)}")
        raise
    
    return order_data


# =====================================================================
# PRODUCT CATALOG FUNCTIONS
# =====================================================================

def _normalize_for_search(s: str) -> str:
    """
    Normalize string for search matching.
    Lowercase, convert non-alphanumeric to space, collapse whitespace.
    
    Args:
        s: String to normalize
        
    Returns:
        Normalized string
    """
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return ' '.join(s.split())


def _collapse_whitespace(s: str) -> str:
    """
    Remove all whitespace, dots, dashes, underscores, and slashes.
    Used for collapsed search matching.
    
    Args:
        s: String to collapse
        
    Returns:
        Collapsed string
    """
    if not s:
        return ""
    return re.sub(r'[\s\.\-_/]+', '', str(s))


def load_product_catalog_to_state(xlsx_bytes: bytes) -> List[Dict[str, Any]]:
    """
    Load product catalog from Excel file into module-level state.
    
    Expected columns (flexible matching):
    - name / Product / Naam (fallback: first column)
    - description / omschrijving / beschrijving / internal notes (fallback: name)
    - list_price (prefer exact), or first column with "prijs"
    - default_code / internal reference / sku (optional)
    
    Args:
        xlsx_bytes: Bytes content of the Excel file
        
    Returns:
        List of product dictionaries loaded into catalog
    """
    global _PRODUCT_CATALOG
    
    try:
        # Use pandas for flexible column detection
        df = pd.read_excel(BytesIO(xlsx_bytes))
        
        # Log original columns
        logger.info(f"Product catalog columns: {list(df.columns)}")
        
        # Normalize column names for matching
        col_lower = {col.lower().strip(): col for col in df.columns}
        
        # Find name column
        name_col = None
        for key in ['name', 'product', 'naam']:
            if key in col_lower:
                name_col = col_lower[key]
                break
        if not name_col:
            # Fallback to first column
            name_col = df.columns[0]
        
        # Find description column
        desc_col = None
        for key in ['description', 'omschrijving', 'beschrijving', 'internal notes']:
            if key in col_lower:
                desc_col = col_lower[key]
                break
        if not desc_col:
            # Fallback to name column
            desc_col = name_col
        
        # Find price column
        price_col = None
        if 'list_price' in col_lower:
            price_col = col_lower['list_price']
        else:
            # Find first column containing "prijs"
            for col in df.columns:
                if 'prijs' in col.lower():
                    price_col = col
                    break
        
        # Find default_code column (optional)
        code_col = None
        for key in ['default_code', 'internal reference', 'sku', 'artikelnr']:
            if key in col_lower:
                code_col = col_lower[key]
                break
        
        logger.info(f"Column mapping: name={name_col}, description={desc_col}, price={price_col}, code={code_col}")
        
        # Build catalog
        catalog = []
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row[name_col]) or not str(row[name_col]).strip():
                continue
            
            name = str(row[name_col]).strip()
            description = str(row[desc_col]).strip() if not pd.isna(row[desc_col]) else name
            
            # Parse price
            price = 0.0
            if price_col and not pd.isna(row[price_col]):
                try:
                    price = float(row[price_col])
                except (ValueError, TypeError):
                    price = 0.0
            
            # Parse code
            code = ""
            if code_col and not pd.isna(row[code_col]):
                code = str(row[code_col]).strip()
            
            # Build search strings
            search_parts = [name, description]
            if code:
                search_parts.append(code)
            
            search_text = " ".join(search_parts)
            search_normalized = _normalize_for_search(search_text)
            search_collapsed = _collapse_whitespace(search_normalized)
            
            catalog.append({
                "name": name,
                "description": description,
                "list_price": price,
                "default_code": code,
                "_search": search_normalized,
                "_search_collapse": search_collapsed
            })
        
        # Update global catalog
        _PRODUCT_CATALOG = catalog
        
        logger.info(f"Loaded {len(catalog)} products into catalog")
        
        return catalog
        
    except Exception as e:
        logger.error(f"Error loading product catalog: {str(e)}")
        raise


def get_product_catalog() -> List[Dict[str, Any]]:
    """
    Get the current product catalog from module-level state.
    
    Returns:
        List of product dictionaries in catalog
    """
    return _PRODUCT_CATALOG


def clear_product_catalog() -> None:
    """
    Clear the product catalog from module-level state.
    """
    global _PRODUCT_CATALOG
    _PRODUCT_CATALOG = []
    logger.info("Product catalog cleared")
